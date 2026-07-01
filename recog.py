import os
import sys
import json
import argparse
import shutil
import pandas as pd
import camelot
from glob import glob

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

from normalizer import process_pdf_tables, STANDARD_COLUMNS


def extract_with_camelot(pdf_path):
    tables = []
    print(f"  Camelot (lattice)...")
    try:
        lattice_tables = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        if len(lattice_tables) > 0:
            print(f"    Найдено {len(lattice_tables)} таблиц(а) методом lattice")
            tables.extend(lattice_tables)
    except Exception as e:
        print(f"    Ошибка lattice: {e}")

    if not tables:
        print(f"  Camelot (stream)...")
        try:
            stream_tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
            if len(stream_tables) > 0:
                print(f"    Найдено {len(stream_tables)} таблиц(а) методом stream")
                tables.extend(stream_tables)
        except Exception as e:
            print(f"    Ошибка stream: {e}")

    return tables


def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    block_size = len(cfg["block_columns"])
    fixed_len = len(cfg["fixed_columns"])
    cfg["_block_size"] = block_size
    cfg["_fixed_len"] = fixed_len

    for i, m in enumerate(cfg["row"].get("meta", [])):
        m.setdefault("value_row", None)

    return cfg


def build_workbook(config, n_suppliers, request_name=""):
    wb = Workbook()
    ws = wb.active
    ws.title = config["sheet_name"]

    n_fixed = config["_fixed_len"]
    block_size = config["_block_size"]
    defaults = config["defaults"]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    font_size = defaults.get("font_size", 11)
    font_name = defaults.get("font_name", "Calibri")
    bold_font = Font(name=font_name, size=font_size, bold=True)
    normal_font = Font(name=font_name, size=font_size)

    for i, col_cfg in enumerate(config["fixed_columns"]):
        cl = get_column_letter(i + 1)
        if col_cfg.get("width"):
            ws.column_dimensions[cl].width = col_cfg["width"]

    for b_idx in range(n_suppliers):
        for col_offset, col_cfg in enumerate(config["block_columns"]):
            col = n_fixed + b_idx * block_size + col_offset + 1
            cl = get_column_letter(col)
            if col_cfg.get("width"):
                ws.column_dimensions[cl].width = col_cfg["width"]
            if col_cfg.get("hidden"):
                ws.column_dimensions[cl].hidden = True

    center_align = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).value = f"Заявка: {request_name}"
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_fixed)
    ws.cell(row=1, column=n_fixed).border = thin_border
    for b_idx in range(n_suppliers):
        sc = n_fixed + b_idx * block_size + 1
        ec = sc + block_size - 1
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        cell = ws.cell(row=1, column=sc)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(row=1, column=ec).border = thin_border

    for i, col_cfg in enumerate(config["fixed_columns"]):
        cell = ws.cell(row=2, column=i + 1)
        cell.value = col_cfg["header"]
        cell.font = bold_font
        cell.border = thin_border
    for b_idx in range(n_suppliers):
        for col_offset, col_cfg in enumerate(config["block_columns"]):
            col = n_fixed + b_idx * block_size + col_offset + 1
            cell = ws.cell(row=2, column=col)
            cell.value = col_cfg["header"]
            cell.font = bold_font
            cell.border = thin_border

    row_cfg = config["row"]
    data_start = row_cfg["data_start"]
    n_data = row_cfg["data_rows"]
    for r in range(data_start, data_start + n_data):
        for c in range(1, n_fixed + n_suppliers * block_size + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border

    meta_labels = [m["label"] for m in row_cfg["meta"]]
    meta_start = data_start + n_data
    for i, label in enumerate(meta_labels):
        row = meta_start + i
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_fixed)
        cell = ws.cell(row=row, column=1)
        cell.value = label
        cell.font = normal_font
        cell.border = thin_border
        for b_idx in range(n_suppliers):
            sc = n_fixed + b_idx * block_size + 1
            ec = sc + block_size - 1
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
            ws.cell(row=row, column=sc).border = thin_border

    total_row = meta_start + len(meta_labels)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=n_fixed)
    ws.cell(row=total_row, column=1).border = thin_border

    for b_idx in range(n_suppliers):
        sc = n_fixed + b_idx * block_size + 1
        if block_size >= 4:
            ws.merge_cells(start_row=total_row, start_column=sc,
                           end_row=total_row, end_column=sc + 3)
        ws.cell(row=total_row, column=sc).border = thin_border
        cell = ws.cell(row=total_row, column=sc + block_size - 1)
        cell.value = "Сумма"
        cell.font = bold_font
        cell.border = thin_border

    wrap_align = Alignment(wrap_text=True, vertical="center")
    for r in range(2, total_row + 1):
        for c in range(1, n_fixed + n_suppliers * block_size + 1):
            ws.cell(row=r, column=c).alignment = wrap_align

    return wb, ws, data_start, n_data, meta_start, total_row


def _copy_style(src, dst):
    try:
        dst.font = copy(src.font)
    except Exception:
        pass
    try:
        dst.fill = copy(src.fill)
    except Exception:
        pass
    try:
        dst.border = copy(src.border)
    except Exception:
        pass
    try:
        dst.alignment = copy(src.alignment)
    except Exception:
        pass
    try:
        dst.number_format = src.number_format
    except Exception:
        pass


def _copy_block_formatting(ws, dst_start, src_start, max_row, block_size):
    for row in range(1, max_row + 1):
        for i in range(block_size):
            _copy_style(ws.cell(row=row, column=src_start + i),
                        ws.cell(row=row, column=dst_start + i))


def _auto_fit_block_columns(ws, block_start, end_row, block_size):
    for offset in range(1, block_size):
        if offset >= 3:
            continue
        col = block_start + offset
        cl = get_column_letter(col)
        rows_to_check = [2] + list(range(3, end_row + 1))
        if offset == block_size - 1:
            rows_to_check.append(end_row + 1)
        max_len = 0
        for r in rows_to_check:
            val = ws.cell(row=r, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[cl].width = min(max_len + 2, 40)


def _find_or_create_block(ws, existing_blocks, config):
    for b in existing_blocks:
        if not b["name"]:
            b["name"] = "PLACEHOLDER"
            return b

    block_size = config["_block_size"]
    n_fixed = config["_fixed_len"]

    if existing_blocks:
        last = max(b["end"] for b in existing_blocks)
        start = last + 1
    else:
        start = n_fixed + 1
    end = start + block_size - 1

    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    bold_font = Font(name=config["defaults"].get("font_name", "Calibri"),
                     size=config["defaults"].get("font_size", 11), bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, col_cfg in enumerate(config["block_columns"]):
        cell = ws.cell(row=2, column=start + i)
        cell.value = col_cfg["header"]
        cell.font = bold_font
        cell.border = thin_border

    wrap_align = Alignment(wrap_text=True, vertical="center")
    first_col = start
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=first_col).alignment = wrap_align

    for col_offset, col_cfg in enumerate(config["block_columns"]):
        cl = get_column_letter(start + col_offset)
        if col_cfg.get("width"):
            ws.column_dimensions[cl].width = col_cfg["width"]
        if col_cfg.get("hidden"):
            ws.column_dimensions[cl].hidden = True

    block = {"start": start, "end": end, "name": "PLACEHOLDER"}
    existing_blocks.append(block)
    existing_blocks.sort(key=lambda b: b["start"])
    return block


def _to_num(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("nan", "None", ""):
        return None
    try:
        return float(s.replace(",", "."))
    except (ValueError, TypeError):
        return None


def _fill_data_row(ws, row_idx, block_start, row_data, bez_nds):
    ws.cell(row=row_idx, column=block_start).value = row_data.get("Товар")
    ws.cell(row=row_idx, column=block_start + 1).value = _to_num(row_data.get("Кол-во"))
    ws.cell(row=row_idx, column=block_start + 2).value = row_data.get("Ед. изм")

    sum_col = get_column_letter(block_start + 5)
    qty_col = get_column_letter(block_start + 1)
    ws.cell(row=row_idx, column=block_start + 3).value = (
        f'=IFERROR({sum_col}{row_idx}/{qty_col}{row_idx},"")'
    )

    ws.cell(row=row_idx, column=block_start + 4).value = _to_num(bez_nds)
    ws.cell(row=row_idx, column=block_start + 5).value = _to_num(row_data.get("Сумма"))

    for off in (3, 4, 5):
        ws.cell(row=row_idx, column=block_start + off).number_format = "#,##0.00"


def fill_template(pdf_data_list, target_dir, script_dir, output_path=None,
                  block_names=None, request_name="", request_items=None):
    config = load_config(os.path.join(script_dir, "config.json"))

    n_fixed = config["_fixed_len"]
    block_size = config["_block_size"]
    bold_font = Font(name=config["defaults"].get("font_name", "Calibri"),
                     size=config["defaults"].get("font_size", 11), bold=True)

    if output_path is None:
        folder_basename = os.path.basename(target_dir)
        parts = folder_basename.split()
        first_two = " ".join(parts[:2]) if len(parts) >= 2 else (parts[0] if parts else "unknown")
        output_name = f"конкурент {first_two}.xlsx"
        output_path = os.path.join(target_dir, output_name)
    else:
        output_name = os.path.basename(output_path)

    wb, ws, data_start, n_data_rows, meta_start, total_row = \
        build_workbook(config, len(pdf_data_list), request_name)

    existing_blocks = []
    for b_idx in range(len(pdf_data_list)):
        sc = n_fixed + b_idx * block_size + 1
        ec = sc + block_size - 1
        existing_blocks.append({"start": sc, "end": ec, "name": ""})

    data_end = meta_start - 1
    ref_block_start = existing_blocks[0]["start"] if existing_blocks else None
    original_block_count = len(existing_blocks)

    if request_items:
        for i, (name, qty) in enumerate(request_items):
            row_idx = data_start + i
            if row_idx > data_end:
                ws.insert_rows(row_idx, 1)
                data_end += 1
                total_row += 1
                for b in existing_blocks:
                    for col_offset in range(block_size):
                        _copy_style(
                            ws.cell(row=data_start, column=b["start"] + col_offset),
                            ws.cell(row=row_idx, column=b["start"] + col_offset),
                        )
                for col in range(1, n_fixed + 1):
                    _copy_style(ws.cell(row=data_start, column=col),
                                ws.cell(row=row_idx, column=col))
            ws.cell(row=row_idx, column=1, value=i + 1)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=qty)

    for df, orig_headers, bez_nds in pdf_data_list:
        if df.shape[0] == 0:
            continue

        filename = df.columns.get_level_values(0)[0]

        already_exists = any(
            b["name"] and filename in b["name"]
            for b in existing_blocks
        )
        if already_exists:
            continue

        block = _find_or_create_block(ws, existing_blocks, config)
        effective_name = (block_names or {}).get(filename, filename)
        block["name"] = effective_name

        cell = ws.cell(row=1, column=block["start"])
        cell.value = effective_name

        max_data_rows = data_end - data_start + 1
        for i in range(min(len(df), max_data_rows)):
            row_idx = data_start + i
            row_data = {col: df.iloc[i][(filename, col)] for col in STANDARD_COLUMNS}
            bv = bez_nds[i] if i < len(bez_nds) else None
            _fill_data_row(ws, row_idx, block["start"], row_data, bv)

        if len(df) > max_data_rows:
            extra = len(df) - max_data_rows
            ws.insert_rows(data_end + 1, extra)

            for i in range(max_data_rows, len(df)):
                row_idx = data_start + i
                row_data = {col: df.iloc[i][(filename, col)] for col in STANDARD_COLUMNS}
                bv = bez_nds[i] if i < len(bez_nds) else None
                _fill_data_row(ws, row_idx, block["start"], row_data, bv)

            first_meta_row_actual = meta_start
            total_row += extra
            data_end += extra
            meta_start += extra

            for r in range(data_end - extra + 1, data_end + 1):
                for b in existing_blocks:
                    for col_offset in range(block_size):
                        _copy_style(
                            ws.cell(row=data_start, column=b["start"] + col_offset),
                            ws.cell(row=r, column=b["start"] + col_offset),
                        )
                for col in range(1, n_fixed + 1):
                    _copy_style(ws.cell(row=data_start, column=col),
                                ws.cell(row=r, column=col))

    for b in existing_blocks:
        ws.merge_cells(start_row=total_row, start_column=b["start"],
                       end_row=total_row, end_column=b["start"] + 3)

        for col_offset, col_cfg in enumerate(config["block_columns"]):
            if col_cfg.get("hidden"):
                cl = get_column_letter(b["start"] + col_offset)
                ws.column_dimensions[cl].hidden = True

        ws.cell(row=total_row, column=b["start"] + block_size - 1).value = "Сумма"

        total_col_letter = get_column_letter(b["start"] + block_size - 1)
        sum_cell = ws.cell(row=total_row + 1, column=b["start"] + block_size - 1)
        sum_cell.value = f"=SUM({total_col_letter}{data_start}:{total_col_letter}{data_end})"
        sum_cell.font = bold_font
        sum_cell.number_format = "#,##0.00"

    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=n_fixed)

    new_blocks = existing_blocks[original_block_count:] if ref_block_start else []
    for b in new_blocks:
        _copy_block_formatting(ws, b["start"], ref_block_start, total_row + 1, block_size)

    stale_ranges = list(ws.merged_cells.ranges)
    for mc in stale_ranges:
        if data_start <= mc.min_row <= data_end:
            ws.merged_cells.remove(mc)

    meta_start_actual = meta_start
    meta_end_actual = total_row - 1
    for mr in range(meta_start_actual, meta_end_actual + 1):
        ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=n_fixed)
        for b in existing_blocks:
            ws.merge_cells(start_row=mr, start_column=b["start"],
                           end_row=mr, end_column=b["end"])

    for b in existing_blocks:
        _auto_fit_block_columns(ws, b["start"], data_end, block_size)

    wb.save(output_path)
    print(f"\nГотово! Конкурентная таблица сохранена как '{output_name}'")
    return output_path


def process_pdf_file(pdf_path, file_data_list):
    print(f"\nОбработка: {os.path.basename(pdf_path)}")
    tables = extract_with_camelot(pdf_path)
    filename = os.path.basename(pdf_path)

    if tables:
        df, orig, bez_nds = process_pdf_tables(tables, filename)
        n_rows = len(df)
        if n_rows > 0:
            print(f"    Извлечено строк: {n_rows}")
        else:
            print(f"    Таблица не найдена (пустой результат)")
        file_data_list.append((df, orig, bez_nds))
    else:
        print(f"    Не удалось извлечь таблицы")
        empty = pd.DataFrame(columns=pd.MultiIndex.from_product([[filename], STANDARD_COLUMNS]))
        orig = {col: col for col in STANDARD_COLUMNS}
        file_data_list.append((empty, orig, []))


def main():
    parser = argparse.ArgumentParser(
        description="Извлечение таблиц из всех PDF-файлов в папке и заполнение конкурентной таблицы."
    )
    parser.add_argument("folder_path", help="Путь к папке с PDF-файлами")
    args = parser.parse_args()

    if not os.path.isdir(args.folder_path):
        print(f"Ошибка: Папка '{args.folder_path}' не найдена.")
        return

    pdf_files = sorted(glob(os.path.join(args.folder_path, "*.[pP][dD][fF]")))
    if not pdf_files:
        print(f"В папке '{args.folder_path}' не найдено PDF-файлов.")
        return

    print(f"Найдено PDF-файлов: {len(pdf_files)}")
    file_data_list = []

    for pdf_path in pdf_files:
        process_pdf_file(pdf_path, file_data_list)

    if not file_data_list:
        print("Не удалось извлечь ни одной таблицы.")
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    fill_template(file_data_list, args.folder_path, script_dir)


if __name__ == "__main__":
    main()
